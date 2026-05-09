"""
نظام حضور وانصراف v6 - Backend كامل
التعديلات الجديدة:
- تصدير جميع التقارير إلى Excel
- إضافة المكافآت الإضافية والخصومات الإضافية في ملخص الرواتب
- فلترة التسويات اليومية حسب الموظف + تصدير Excel
- تعديل السلف + البحث حسب الموظف والفترة + تصدير Excel
- إصلاحات وتحسينات عامة
"""
from fastapi import FastAPI, Depends, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
import asyncpg, os, math, jwt, bcrypt, io
from datetime import datetime, date, timedelta, time as dt_time
from pydantic import BaseModel
from typing import Optional, List
from contextlib import asynccontextmanager
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()
DATABASE_URL = os.getenv("DATABASE_URL","postgresql://postgres:password@localhost:5432/attendance_db")
JWT_SECRET   = os.getenv("JWT_SECRET","super-secret-change-me-in-production-32chars!!")
JWT_ALGO     = "HS256"

db_pool = None

@asynccontextmanager
async def lifespan(app: FastAPI):
    global db_pool
    db_pool = await asyncpg.create_pool(DATABASE_URL, min_size=2, max_size=15)
    print("✅ DB connected")
    yield
    await db_pool.close()

app = FastAPI(title="نظام الحضور v6", version="6.0.0", lifespan=lifespan)

# ── PORT ديناميكي للـ Hosting ──
PORT = int(os.getenv("PORT", "8000"))

# ── CORS مفتوح بالكامل ──
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["GET","POST","PUT","DELETE","OPTIONS","PATCH"],
    allow_headers=["*"],
    expose_headers=["*"],
)

security = HTTPBearer()

# ── Models ──
class LoginReq(BaseModel):
    username: str; password: str; subdomain: str

class TenantUpdateReq(BaseModel):
    name: Optional[str]=None; address: Optional[str]=None; phone: Optional[str]=None
    email: Optional[str]=None; tax_number: Optional[str]=None
    office_latitude: Optional[float]=None; office_longitude: Optional[float]=None
    geofence_radius_meters: Optional[int]=None; work_start_time: Optional[str]=None
    work_end_time: Optional[str]=None; late_threshold_minutes: Optional[int]=None
    currency: Optional[str]=None

class DeptReq(BaseModel):
    name: str; description: Optional[str]=None; manager_id: Optional[str]=None
    work_start_time: Optional[str]=None; work_end_time: Optional[str]=None
    late_threshold_minutes: Optional[int]=None; is_flexible: bool=False

class CreateUserReq(BaseModel):
    full_name: str; username: str; password: str
    role: str="employee"; department_id: Optional[str]=None
    job_title: Optional[str]=None; employee_code: Optional[str]=None
    phone: Optional[str]=None; national_id: Optional[str]=None
    hire_date: Optional[date]=None
    hourly_rate: float=0.0; base_salary: float=0.0
    salary_type: str="hourly"

class UpdateUserReq(BaseModel):
    full_name: Optional[str]=None; role: Optional[str]=None
    department_id: Optional[str]=None; job_title: Optional[str]=None
    employee_code: Optional[str]=None; phone: Optional[str]=None
    national_id: Optional[str]=None; hire_date: Optional[date]=None
    hourly_rate: Optional[float]=None; base_salary: Optional[float]=None
    salary_type: Optional[str]=None; is_active: Optional[bool]=None
    password: Optional[str]=None

class CheckInReq(BaseModel):
    latitude: float; longitude: float; accuracy: Optional[float]=None
    is_remote: bool=False; notes: Optional[str]=None

class CheckOutReq(BaseModel):
    latitude: float; longitude: float; accuracy: Optional[float]=None

class DailyAdjReq(BaseModel):
    user_id: str; adjustment_date: date
    bonus: float=0.0; deduction: float=0.0; reason: Optional[str]=None

class UpdateAdjReq(BaseModel):
    bonus: Optional[float]=None
    deduction: Optional[float]=None
    reason: Optional[str]=None

class AdvanceReq(BaseModel):
    user_id: str; amount: float; advance_date: date
    reason: Optional[str]=None; notes: Optional[str]=None

class UpdateAdvanceReq(BaseModel):
    amount: Optional[float]=None
    advance_date: Optional[date]=None
    reason: Optional[str]=None
    notes: Optional[str]=None
    status: Optional[str]=None

class AdvanceStatusReq(BaseModel):
    status: str

class SalaryCalcReq(BaseModel):
    user_id: str; date_from: date; date_to: date
    extra_bonus: float=0.0; extra_deduction: float=0.0
    notes: Optional[str]=None; report_name: Optional[str]=None

class ManualAttendanceReq(BaseModel):
    user_id: str
    attendance_date: date
    check_in_time: Optional[str]=None
    check_out_time: Optional[str]=None
    status: str="present"
    notes: Optional[str]=None
    is_remote: bool=False

class EditAttendanceReq(BaseModel):
    check_in_time: Optional[str]=None
    check_out_time: Optional[str]=None
    attendance_date: Optional[date]=None
    status: Optional[str]=None
    notes: Optional[str]=None
    is_remote: Optional[bool]=None

class UpdateSalaryReportReq(BaseModel):
    report_name: Optional[str]=None
    notes: Optional[str]=None

# ── Helpers ──
def haversine(lat1,lon1,lat2,lon2):
    R=6371000; p1,p2=math.radians(lat1),math.radians(lat2)
    dp,dl=math.radians(lat2-lat1),math.radians(lon2-lon1)
    a=math.sin(dp/2)**2+math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return R*2*math.atan2(math.sqrt(a),math.sqrt(1-a))

def make_token(uid,tid,role):
    return jwt.encode({"user_id":uid,"tenant_id":tid,"role":role,
        "exp":datetime.utcnow()+timedelta(hours=24)},JWT_SECRET,algorithm=JWT_ALGO)

async def get_user(creds: HTTPAuthorizationCredentials=Depends(security)):
    try: return jwt.decode(creds.credentials,JWT_SECRET,algorithms=[JWT_ALGO])
    except jwt.ExpiredSignatureError: raise HTTPException(401,"انتهت الجلسة")
    except: raise HTTPException(401,"بيانات دخول غير صالحة")

def admin_only(u=Depends(get_user)):
    if u["role"] not in ("admin","manager"): raise HTTPException(403,"غير مصرح")
    return u

def s(v): return str(v) if v else None

# ── Excel Export Helper ──
def create_excel_response(wb, filename):
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"',
        'Access-Control-Expose-Headers': 'Content-Disposition'
    }
    return StreamingResponse(buffer, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)

def style_header(ws, row_num, cols_count):
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(1, cols_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

def auto_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

# ── Health ──
@app.get("/health")
async def health(): return {"status":"ok","time":datetime.now().isoformat()}

# ── AUTH ──
@app.post("/api/auth/login",tags=["Auth"])
async def login(r: LoginReq):
    async with db_pool.acquire() as c:
        tenant=await c.fetchrow("SELECT * FROM tenants WHERE subdomain=$1 AND is_active=TRUE",r.subdomain)
        if not tenant: raise HTTPException(404,"الشركة غير موجودة")
        user=await c.fetchrow(
            """SELECT u.*,d.name as dept_name,d.work_start_time as dept_start,
                      d.late_threshold_minutes as dept_late,d.is_flexible as dept_flexible
               FROM users u LEFT JOIN departments d ON u.department_id=d.id
               WHERE u.tenant_id=$1 AND u.username=$2 AND u.is_active=TRUE""",
            tenant["id"],r.username)
        if not user: raise HTTPException(401,"اسم المستخدم أو كلمة المرور غير صحيحة")
        if not bcrypt.checkpw(r.password.encode(),user["password_hash"].encode()):
            raise HTTPException(401,"اسم المستخدم أو كلمة المرور غير صحيحة")
        await c.execute("UPDATE users SET last_login_at=NOW() WHERE id=$1",user["id"])
        return {
            "access_token":make_token(s(user["id"]),s(tenant["id"]),user["role"]),
            "user":{"id":s(user["id"]),"full_name":user["full_name"],"username":user["username"],
                    "role":user["role"],"job_title":user["job_title"],
                    "hourly_rate":float(user["hourly_rate"] or 0),
                    "base_salary":float(user["base_salary"] or 0),
                    "salary_type":user["salary_type"],
                    "department_id":s(user["department_id"]),"dept_name":user["dept_name"],
                    "dept_flexible":user["dept_flexible"]},
            "tenant":{"id":s(tenant["id"]),"name":tenant["name"],
                      "work_start_time":str(tenant["work_start_time"]),
                      "currency":tenant["currency"]}
        }

# ── TENANT SETTINGS ──
@app.get("/api/settings/tenant",tags=["Settings"])
async def get_tenant(u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        t=await c.fetchrow("SELECT * FROM tenants WHERE id=$1",u["tenant_id"])
        return dict(t)|{"id":s(t["id"]),
            "office_latitude":float(t["office_latitude"]) if t["office_latitude"] else None,
            "office_longitude":float(t["office_longitude"]) if t["office_longitude"] else None,
            "work_start_time":str(t["work_start_time"]) if t["work_start_time"] else None,
            "work_end_time":str(t["work_end_time"]) if t["work_end_time"] else None,
            "created_at":t["created_at"].isoformat()}

@app.put("/api/settings/tenant",tags=["Settings"])
async def update_tenant(r: TenantUpdateReq,u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        data=r.model_dump(exclude_none=True)
        if not data: raise HTTPException(400,"لا توجد بيانات للتحديث")
        fields=[f"{k}=${i+1}" for i,k in enumerate(data.keys())]
        vals=list(data.values())+[u["tenant_id"]]
        await c.execute(f"UPDATE tenants SET {','.join(fields)},updated_at=NOW() WHERE id=${len(vals)}",*vals)
        return {"success":True,"message":"تم حفظ إعدادات الشركة ✅"}

# ── DEPARTMENTS ──
@app.get("/api/departments",tags=["Departments"])
async def list_depts(u=Depends(get_user)):
    async with db_pool.acquire() as c:
        rows=await c.fetch(
            """SELECT d.*,u.full_name as manager_name,
                      (SELECT COUNT(*) FROM users WHERE department_id=d.id AND is_active=TRUE) as emp_count
               FROM departments d LEFT JOIN users u ON d.manager_id=u.id
               WHERE d.tenant_id=$1 AND d.is_active=TRUE ORDER BY d.name""",u["tenant_id"])
        return [dict(r)|{"id":s(r["id"]),"manager_id":s(r["manager_id"]),
            "work_start_time":str(r["work_start_time"]) if r["work_start_time"] else None,
            "work_end_time":str(r["work_end_time"]) if r["work_end_time"] else None,
            "created_at":r["created_at"].isoformat()} for r in rows]

@app.post("/api/departments",tags=["Departments"])
async def create_dept(r: DeptReq,u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        if await c.fetchval("SELECT id FROM departments WHERE tenant_id=$1 AND name=$2 AND is_active=TRUE",u["tenant_id"],r.name):
            raise HTTPException(400,"القسم موجود مسبقاً")
        row=await c.fetchrow(
            "INSERT INTO departments (tenant_id,name,description,manager_id,work_start_time,work_end_time,late_threshold_minutes,is_flexible) VALUES ($1,$2,$3,$4,$5,$6,$7,$8) RETURNING id",
            u["tenant_id"],r.name,r.description,r.manager_id,r.work_start_time,r.work_end_time,r.late_threshold_minutes,r.is_flexible)
        return {"success":True,"message":f"تم إنشاء قسم '{r.name}'","id":s(row["id"])}

@app.put("/api/departments/{did}",tags=["Departments"])
async def update_dept(did:str,r: DeptReq,u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        await c.execute(
            "UPDATE departments SET name=$1,description=$2,manager_id=$3,work_start_time=$4,work_end_time=$5,late_threshold_minutes=$6,is_flexible=$7 WHERE id=$8 AND tenant_id=$9",
            r.name,r.description,r.manager_id,r.work_start_time,r.work_end_time,r.late_threshold_minutes,r.is_flexible,did,u["tenant_id"])
        return {"success":True,"message":"تم تحديث القسم"}

@app.delete("/api/departments/{did}",tags=["Departments"])
async def delete_dept(did:str,u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        cnt=await c.fetchval("SELECT COUNT(*) FROM users WHERE department_id=$1 AND is_active=TRUE",did)
        if cnt>0: raise HTTPException(400,f"القسم يحتوي على {cnt} موظف")
        await c.execute("UPDATE departments SET is_active=FALSE WHERE id=$1 AND tenant_id=$2",did,u["tenant_id"])
        return {"success":True,"message":"تم حذف القسم"}

# ── USERS ──
@app.get("/api/users",tags=["Users"])
async def list_users(u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        rows=await c.fetch(
            "SELECT u.*,d.name as dept_name FROM users u LEFT JOIN departments d ON u.department_id=d.id WHERE u.tenant_id=$1 ORDER BY u.full_name",
            u["tenant_id"])
        return [dict(r)|{"id":s(r["id"]),"department_id":s(r["department_id"]),
            "hourly_rate":float(r["hourly_rate"] or 0),"base_salary":float(r["base_salary"] or 0),
            "hire_date":str(r["hire_date"]) if r["hire_date"] else None,
            "created_at":r["created_at"].isoformat()} for r in rows]

@app.post("/api/users",tags=["Users"])
async def create_user(r: CreateUserReq,u=Depends(admin_only)):
    hashed=bcrypt.hashpw(r.password.encode(),bcrypt.gensalt()).decode()
    async with db_pool.acquire() as c:
        if await c.fetchval("SELECT id FROM users WHERE tenant_id=$1 AND username=$2",u["tenant_id"],r.username):
            raise HTTPException(400,"اسم المستخدم مستخدم مسبقاً")
        row=await c.fetchrow(
            "INSERT INTO users (tenant_id,department_id,full_name,username,password_hash,role,job_title,employee_code,phone,national_id,hire_date,hourly_rate,base_salary,salary_type) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14) RETURNING id",
            u["tenant_id"],r.department_id,r.full_name,r.username,hashed,r.role,r.job_title,r.employee_code,r.phone,r.national_id,r.hire_date,r.hourly_rate,r.base_salary,r.salary_type)
        return {"success":True,"message":f"تم إضافة '{r.full_name}'","id":s(row["id"])}

@app.get("/api/users/{uid}",tags=["Users"])
async def get_user_detail(uid:str,u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        row=await c.fetchrow(
            "SELECT u.*,d.name as dept_name FROM users u LEFT JOIN departments d ON u.department_id=d.id WHERE u.id=$1 AND u.tenant_id=$2",
            uid,u["tenant_id"])
        if not row: raise HTTPException(404,"الموظف غير موجود")
        return dict(row)|{"id":s(row["id"]),"department_id":s(row["department_id"]),
            "hourly_rate":float(row["hourly_rate"] or 0),"base_salary":float(row["base_salary"] or 0),
            "hire_date":str(row["hire_date"]) if row["hire_date"] else None}

@app.put("/api/users/{uid}",tags=["Users"])
async def update_user(uid:str,r: UpdateUserReq,u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        data = r.model_dump(exclude_none=True)
        if 'password' in data and data['password']:
            data['password_hash'] = bcrypt.hashpw(data['password'].encode(), bcrypt.gensalt()).decode()
            del data['password']
        elif 'password' in data:
            del data['password']
        if not data: raise HTTPException(400,"لا توجد بيانات")
        fields=[f"{k}=${i+1}" for i,k in enumerate(data.keys())]
        vals=list(data.values())+[uid]
        await c.execute(f"UPDATE users SET {','.join(fields)} WHERE id=${len(vals)}",*vals)
        return {"success":True,"message":"تم تحديث بيانات الموظف"}

@app.delete("/api/users/{uid}",tags=["Users"])
async def deactivate_user(uid:str,u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        await c.execute("UPDATE users SET is_active=FALSE WHERE id=$1 AND tenant_id=$2",uid,u["tenant_id"])
        return {"success":True,"message":"تم تعطيل حساب الموظف"}

# ── ATTENDANCE ──
@app.post("/api/attendance/check-in",tags=["Attendance"])
async def check_in(r: CheckInReq,u=Depends(get_user)):
    async with db_pool.acquire() as c:
        tenant=await c.fetchrow("SELECT * FROM tenants WHERE id=$1",u["tenant_id"])
        user_row=await c.fetchrow(
            "SELECT u.department_id,d.work_start_time,d.late_threshold_minutes,d.is_flexible FROM users u LEFT JOIN departments d ON u.department_id=d.id WHERE u.id=$1",
            u["user_id"])
        if tenant["office_latitude"] and not r.is_remote:
            dist=haversine(r.latitude,r.longitude,float(tenant["office_latitude"]),float(tenant["office_longitude"]))
            if dist>tenant["geofence_radius_meters"]:
                raise HTTPException(400,f"أنت خارج نطاق المكتب — المسافة: {dist:.0f}م")
        today=date.today()
        last_sess=await c.fetchval(
            "SELECT COALESCE(MAX(session_number),0) FROM attendance WHERE user_id=$1 AND attendance_date=$2",
            u["user_id"],today)
        open_sess=await c.fetchrow(
            "SELECT id FROM attendance WHERE user_id=$1 AND attendance_date=$2 AND check_in_time IS NOT NULL AND check_out_time IS NULL ORDER BY session_number DESC LIMIT 1",
            u["user_id"],today)
        if open_sess: raise HTTPException(400,"لديك جلسة مفتوحة — سجّل الانصراف أولاً")
        is_flex=user_row["is_flexible"] if user_row else False
        dept_start=user_row["work_start_time"] if user_row else None
        dept_late=user_row["late_threshold_minutes"] if user_row else None
        if is_flex or (dept_start is None and dept_late is None):
            status="flexible"; late_min=0
        else:
            start_time=dept_start or tenant["work_start_time"]
            threshold=dept_late if dept_late is not None else tenant["late_threshold_minutes"]
            now=datetime.now()
            work_start=datetime.combine(today,start_time)
            late_min=max(0,int((now-work_start).total_seconds()/60))
            status="late" if late_min>threshold else "present"
        row=await c.fetchrow(
            "INSERT INTO attendance (tenant_id,user_id,attendance_date,check_in_time,check_in_latitude,check_in_longitude,status,late_minutes,is_remote,notes,session_number) VALUES ($1,$2,$3,NOW(),$4,$5,$6,$7,$8,$9,$10) RETURNING id,check_in_time,status,late_minutes,session_number",
            u["tenant_id"],u["user_id"],today,r.latitude,r.longitude,status,late_min,r.is_remote,r.notes,last_sess+1)
        msgs={"present":"✅ تم تسجيل حضورك","late":f"⚠️ تم تسجيل حضورك — تأخير {late_min} دقيقة","flexible":"✅ تم تسجيل حضورك (دوام مرن)"}
        return {"success":True,"message":msgs.get(status,"✅ تم التسجيل"),
                "session_number":row["session_number"],"check_in_time":row["check_in_time"].isoformat(),
                "status":row["status"],"late_minutes":row["late_minutes"]}

@app.post("/api/attendance/check-out",tags=["Attendance"])
async def check_out(r: CheckOutReq,u=Depends(get_user)):
    async with db_pool.acquire() as c:
        sess=await c.fetchrow(
            "SELECT id FROM attendance WHERE user_id=$1 AND attendance_date=$2 AND check_in_time IS NOT NULL AND check_out_time IS NULL ORDER BY session_number DESC LIMIT 1",
            u["user_id"],date.today())
        if not sess: raise HTTPException(400,"لا توجد جلسة مفتوحة")
        row=await c.fetchrow(
            "UPDATE attendance SET check_out_time=NOW(),check_out_latitude=$1,check_out_longitude=$2 WHERE id=$3 RETURNING check_out_time,work_hours,session_number",
            r.latitude,r.longitude,sess["id"])
        return {"success":True,"message":"👋 تم تسجيل الانصراف",
                "session_number":row["session_number"],
                "check_out_time":row["check_out_time"].isoformat(),
                "work_hours":float(row["work_hours"]) if row["work_hours"] else 0}

@app.get("/api/attendance/today",tags=["Attendance"])
async def today_status(u=Depends(get_user)):
    async with db_pool.acquire() as c:
        rows=await c.fetch(
            "SELECT * FROM attendance WHERE user_id=$1 AND attendance_date=$2 ORDER BY session_number",
            u["user_id"],date.today())
        sessions=[{"id":s(r["id"]),"session_number":r["session_number"],
            "check_in_time":r["check_in_time"].isoformat() if r["check_in_time"] else None,
            "check_out_time":r["check_out_time"].isoformat() if r["check_out_time"] else None,
            "status":r["status"],"late_minutes":r["late_minutes"],
            "work_hours":float(r["work_hours"]) if r["work_hours"] else None,
            "is_manual":r["is_manual"]} for r in rows]
        has_open=any(not r["check_out_time"] and r["check_in_time"] for r in rows)
        total_hours=sum(float(r["work_hours"] or 0) for r in rows)
        return {"sessions":sessions,"has_open_session":has_open,
                "total_sessions":len(sessions),"total_hours":round(total_hours,2)}

# ── MANUAL ATTENDANCE (Admin) ──
@app.get("/api/admin/attendance/{user_id}/{att_date}",tags=["Manual Attendance"])
async def get_user_attendance_day(user_id: str, att_date: date, u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        rows = await c.fetch(
            """SELECT a.*, u.full_name 
               FROM attendance a 
               JOIN users u ON a.user_id = u.id 
               WHERE a.tenant_id=$1 AND a.user_id=$2 AND a.attendance_date=$3 
               ORDER BY a.session_number""",
            u["tenant_id"], user_id, att_date)
        return [{
            "id": s(r["id"]),
            "session_number": r["session_number"],
            "attendance_date": str(r["attendance_date"]),
            "check_in_time": r["check_in_time"].isoformat() if r["check_in_time"] else None,
            "check_out_time": r["check_out_time"].isoformat() if r["check_out_time"] else None,
            "status": r["status"],
            "late_minutes": r["late_minutes"],
            "work_hours": float(r["work_hours"]) if r["work_hours"] else None,
            "is_remote": r["is_remote"],
            "is_manual": r["is_manual"],
            "notes": r["notes"],
            "full_name": r["full_name"]
        } for r in rows]

@app.post("/api/admin/attendance/manual",tags=["Manual Attendance"])
async def add_manual_attendance(r: ManualAttendanceReq, u=Depends(admin_only)):
    """إضافة حضور يدوي — يدعم تاريخ مخصص (ماضٍ أو حاضر)"""
    async with db_pool.acquire() as c:
        user = await c.fetchrow(
            "SELECT id, full_name FROM users WHERE id=$1 AND tenant_id=$2 AND is_active=TRUE",
            r.user_id, u["tenant_id"])
        if not user:
            raise HTTPException(404, "الموظف غير موجود")

        check_in_dt = None
        check_out_dt = None
        att_date = r.attendance_date

        if r.check_in_time:
            try:
                time_parts = r.check_in_time.split(':')
                check_in_dt = datetime.combine(att_date,
                    dt_time(int(time_parts[0]), int(time_parts[1]),
                    int(time_parts[2]) if len(time_parts) > 2 else 0))
            except:
                raise HTTPException(400, "صيغة وقت الدخول غير صحيحة (HH:MM:SS)")

        if r.check_out_time:
            try:
                time_parts = r.check_out_time.split(':')
                check_out_dt = datetime.combine(att_date,
                    dt_time(int(time_parts[0]), int(time_parts[1]),
                    int(time_parts[2]) if len(time_parts) > 2 else 0))
            except:
                raise HTTPException(400, "صيغة وقت الخروج غير صحيحة (HH:MM:SS)")

        late_min = 0
        tenant = await c.fetchrow("SELECT work_start_time, late_threshold_minutes FROM tenants WHERE id=$1", u["tenant_id"])
        if check_in_dt and r.status not in ('flexible', 'half_day'):
            work_start = datetime.combine(att_date, tenant["work_start_time"])
            if check_in_dt > work_start:
                late_min = int((check_in_dt - work_start).total_seconds() / 60)

        last_sess = await c.fetchval(
            "SELECT COALESCE(MAX(session_number),0) FROM attendance WHERE user_id=$1 AND attendance_date=$2",
            r.user_id, att_date)

        row = await c.fetchrow(
            """INSERT INTO attendance 
               (tenant_id, user_id, attendance_date, check_in_time, check_out_time, 
                session_number, status, late_minutes, is_remote, notes, is_manual, created_by)
               VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,TRUE,$11)
               RETURNING id, session_number""",
            u["tenant_id"], r.user_id, att_date, check_in_dt, check_out_dt,
            last_sess + 1, r.status, late_min, r.is_remote, r.notes, u["user_id"])

        return {
            "success": True,
            "message": f"تم تسجيل حضور يدوي لـ {user['full_name']} بتاريخ {att_date}",
            "id": s(row["id"]),
            "session_number": row["session_number"]
        }

@app.put("/api/admin/attendance/{att_id}",tags=["Manual Attendance"])
async def edit_attendance(att_id: str, r: EditAttendanceReq, u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        att = await c.fetchrow(
            "SELECT * FROM attendance WHERE id=$1 AND tenant_id=$2",
            att_id, u["tenant_id"])
        if not att:
            raise HTTPException(404, "سجل الحضور غير موجود")

        ref_date = r.attendance_date if r.attendance_date else att["attendance_date"]
        updates = {}

        if r.attendance_date is not None:
            updates["attendance_date"] = r.attendance_date

        if r.check_in_time is not None:
            try:
                tp = r.check_in_time.split(':')
                updates["check_in_time"] = datetime.combine(ref_date,
                    dt_time(int(tp[0]), int(tp[1]), int(tp[2]) if len(tp) > 2 else 0))
            except:
                raise HTTPException(400, "صيغة وقت الدخول غير صحيحة")

        if r.check_out_time is not None:
            try:
                tp = r.check_out_time.split(':')
                updates["check_out_time"] = datetime.combine(ref_date,
                    dt_time(int(tp[0]), int(tp[1]), int(tp[2]) if len(tp) > 2 else 0))
            except:
                raise HTTPException(400, "صيغة وقت الخروج غير صحيحة")

        if r.status is not None: updates["status"] = r.status
        if r.notes is not None: updates["notes"] = r.notes
        if r.is_remote is not None: updates["is_remote"] = r.is_remote
        updates["is_manual"] = True

        if not updates: raise HTTPException(400, "لا توجد بيانات للتحديث")

        fields = [f"{k}=${i+1}" for i, k in enumerate(updates.keys())]
        vals = list(updates.values()) + [att_id]
        await c.execute(f"UPDATE attendance SET {','.join(fields)} WHERE id=${len(vals)}", *vals)
        return {"success": True, "message": "تم تعديل سجل الحضور"}

@app.delete("/api/admin/attendance/{att_id}",tags=["Manual Attendance"])
async def delete_attendance(att_id: str, u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        result = await c.execute(
            "DELETE FROM attendance WHERE id=$1 AND tenant_id=$2",
            att_id, u["tenant_id"])
        if result == "DELETE 0":
            raise HTTPException(404, "سجل الحضور غير موجود")
        return {"success": True, "message": "تم حذف سجل الحضور"}

# ── ADJUSTMENTS ── (مع تعديل وحذف)
@app.post("/api/adjustments",tags=["Adjustments"])
async def create_adj(r: DailyAdjReq,u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        row = await c.fetchrow(
            """INSERT INTO daily_adjustments (tenant_id,user_id,adjustment_date,bonus,deduction,reason,created_by) 
               VALUES ($1,$2,$3,$4,$5,$6,$7) 
               ON CONFLICT (tenant_id,user_id,adjustment_date) 
               DO UPDATE SET bonus=$4,deduction=$5,reason=$6,created_by=$7
               RETURNING id""",
            u["tenant_id"],r.user_id,r.adjustment_date,r.bonus,r.deduction,r.reason,u["user_id"])
        return {"success":True,"message":"تم حفظ التسوية","id":s(row["id"])}

@app.put("/api/adjustments/{adj_id}",tags=["Adjustments"])
async def update_adj(adj_id: str, r: UpdateAdjReq, u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        adj = await c.fetchrow(
            "SELECT * FROM daily_adjustments WHERE id=$1 AND tenant_id=$2",
            adj_id, u["tenant_id"])
        if not adj: raise HTTPException(404, "التسوية غير موجودة")
        data = r.model_dump(exclude_none=True)
        if not data: raise HTTPException(400, "لا توجد بيانات")
        fields = [f"{k}=${i+1}" for i, k in enumerate(data.keys())]
        vals = list(data.values()) + [adj_id]
        await c.execute(f"UPDATE daily_adjustments SET {','.join(fields)} WHERE id=${len(vals)}", *vals)
        return {"success": True, "message": "تم تحديث التسوية"}

@app.delete("/api/adjustments/{adj_id}",tags=["Adjustments"])
async def delete_adj(adj_id: str, u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        result = await c.execute(
            "DELETE FROM daily_adjustments WHERE id=$1 AND tenant_id=$2",
            adj_id, u["tenant_id"])
        if result == "DELETE 0": raise HTTPException(404, "التسوية غير موجودة")
        return {"success": True, "message": "تم حذف التسوية"}

@app.get("/api/adjustments/date/{adj_date}",tags=["Adjustments"])
async def day_adjs(adj_date:date,u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        rows=await c.fetch(
            "SELECT da.*,u.full_name,u.employee_code FROM daily_adjustments da JOIN users u ON da.user_id=u.id WHERE da.tenant_id=$1 AND da.adjustment_date=$2 ORDER BY u.full_name",
            u["tenant_id"],adj_date)
        return [dict(r)|{"id":s(r["id"]),"user_id":s(r["user_id"]),
            "bonus":float(r["bonus"]),"deduction":float(r["deduction"])} for r in rows]

@app.get("/api/adjustments/range",tags=["Adjustments"])
async def range_adjs(
    date_from: date = Query(...), 
    date_to: date = Query(...), 
    user_id: Optional[str] = Query(None),
    u=Depends(admin_only)
):
    """عرض التسويات لفترة زمنية من/إلى مع فلترة حسب الموظف"""
    async with db_pool.acquire() as c:
        query = """SELECT da.*,u.full_name,u.employee_code 
               FROM daily_adjustments da JOIN users u ON da.user_id=u.id 
               WHERE da.tenant_id=$1 AND da.adjustment_date BETWEEN $2 AND $3"""
        params = [u["tenant_id"], date_from, date_to]
        if user_id:
            query += f" AND da.user_id=${len(params)+1}"
            params.append(user_id)
        query += " ORDER BY da.adjustment_date DESC, u.full_name"
        rows = await c.fetch(query, *params)
        return [dict(r)|{"id":s(r["id"]),"user_id":s(r["user_id"]),
            "bonus":float(r["bonus"]),"deduction":float(r["deduction"]),
            "adjustment_date":str(r["adjustment_date"])} for r in rows]

# ── Export Adjustments to Excel ──
@app.get("/api/adjustments/range/export",tags=["Adjustments"])
async def export_range_adjs(
    date_from: date = Query(...), 
    date_to: date = Query(...), 
    user_id: Optional[str] = Query(None),
    u=Depends(admin_only)
):
    """تصدير التسويات لفترة زمنية إلى Excel"""
    async with db_pool.acquire() as c:
        query = """SELECT da.*,u.full_name,u.employee_code 
               FROM daily_adjustments da JOIN users u ON da.user_id=u.id 
               WHERE da.tenant_id=$1 AND da.adjustment_date BETWEEN $2 AND $3"""
        params = [u["tenant_id"], date_from, date_to]
        if user_id:
            query += f" AND da.user_id=${len(params)+1}"
            params.append(user_id)
        query += " ORDER BY da.adjustment_date DESC, u.full_name"
        rows = await c.fetch(query, *params)

        wb = Workbook()
        ws = wb.active
        ws.title = "التسويات اليومية"
        ws.sheet_view.rightToLeft = True

        headers = ["التاريخ", "كود الموظف", "الاسم", "مكافأة", "خصم", "السبب"]
        ws.append(headers)
        style_header(ws, 1, len(headers))

        for r in rows:
            ws.append([
                str(r["adjustment_date"]),
                r["employee_code"] or "",
                r["full_name"],
                float(r["bonus"] or 0),
                float(r["deduction"] or 0),
                r["reason"] or ""
            ])

        auto_width(ws)
        filename = f"adjustments_{date_from}_{date_to}.xlsx"
        return create_excel_response(wb, filename)

# ── ADVANCES ── (مع تعديل + بحث + تصدير)
@app.get("/api/advances",tags=["Advances"])
async def list_advances(
    u=Depends(admin_only),
    status:Optional[str]=None,
    user_id: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None
):
    async with db_pool.acquire() as c:
        q="SELECT a.*,u.full_name,u.employee_code FROM advances a JOIN users u ON a.user_id=u.id WHERE a.tenant_id=$1"
        params=[u["tenant_id"]]
        if status: 
            q+=f" AND a.status=${len(params)+1}"; params.append(status)
        if user_id: 
            q+=f" AND a.user_id=${len(params)+1}"; params.append(user_id)
        if date_from: 
            q+=f" AND a.advance_date>=${len(params)+1}"; params.append(date_from)
        if date_to: 
            q+=f" AND a.advance_date<=${len(params)+1}"; params.append(date_to)
        q+=" ORDER BY a.advance_date DESC"
        rows=await c.fetch(q,*params)
        return [dict(r)|{"id":s(r["id"]),"user_id":s(r["user_id"]),
            "amount":float(r["amount"]),"advance_date":str(r["advance_date"])} for r in rows]

@app.post("/api/advances",tags=["Advances"])
async def create_advance(r: AdvanceReq,u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        if r.amount<=0: raise HTTPException(400,"المبلغ يجب أن يكون أكبر من صفر")
        row=await c.fetchrow(
            "INSERT INTO advances (tenant_id,user_id,amount,advance_date,reason,notes,status,approved_by,approved_at) VALUES ($1,$2,$3,$4,$5,$6,'approved',$7,NOW()) RETURNING id",
            u["tenant_id"],r.user_id,r.amount,r.advance_date,r.reason,r.notes,u["user_id"])
        return {"success":True,"message":f"تم تسجيل سلفة {r.amount}","id":s(row["id"])}

@app.put("/api/advances/{aid}",tags=["Advances"])
async def update_advance(aid:str,r: UpdateAdvanceReq,u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        adv = await c.fetchrow("SELECT * FROM advances WHERE id=$1 AND tenant_id=$2", aid, u["tenant_id"])
        if not adv: raise HTTPException(404, "السلفة غير موجودة")
        data = r.model_dump(exclude_none=True)
        if not data: raise HTTPException(400, "لا توجد بيانات")
        if 'status' in data:
            data['approved_by'] = u["user_id"]
            data['approved_at'] = datetime.now()
        fields = [f"{k}=${i+1}" for i, k in enumerate(data.keys())]
        vals = list(data.values()) + [aid]
        await c.execute(f"UPDATE advances SET {','.join(fields)} WHERE id=${len(vals)}", *vals)
        return {"success":True,"message":"تم تحديث السلفة"}

@app.delete("/api/advances/{aid}",tags=["Advances"])
async def delete_advance(aid:str,u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        await c.execute("DELETE FROM advances WHERE id=$1 AND tenant_id=$2",aid,u["tenant_id"])
        return {"success":True,"message":"تم حذف السلفة"}

# ── Export Advances to Excel ──
@app.get("/api/advances/export",tags=["Advances"])
async def export_advances(
    u=Depends(admin_only),
    status:Optional[str]=None,
    user_id: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None
):
    """تصدير السلف إلى Excel"""
    async with db_pool.acquire() as c:
        q="SELECT a.*,u.full_name,u.employee_code FROM advances a JOIN users u ON a.user_id=u.id WHERE a.tenant_id=$1"
        params=[u["tenant_id"]]
        if status: q+=f" AND a.status=${len(params)+1}"; params.append(status)
        if user_id: q+=f" AND a.user_id=${len(params)+1}"; params.append(user_id)
        if date_from: q+=f" AND a.advance_date>=${len(params)+1}"; params.append(date_from)
        if date_to: q+=f" AND a.advance_date<=${len(params)+1}"; params.append(date_to)
        q+=" ORDER BY a.advance_date DESC"
        rows=await c.fetch(q,*params)

        wb = Workbook()
        ws = wb.active
        ws.title = "السلف"
        ws.sheet_view.rightToLeft = True

        headers = ["التاريخ", "كود الموظف", "الاسم", "المبلغ", "الحالة", "السبب", "ملاحظات"]
        ws.append(headers)
        style_header(ws, 1, len(headers))

        status_map = {'approved':'معتمدة','pending':'معلقة','rejected':'مرفوضة','paid':'مدفوعة'}
        for r in rows:
            ws.append([
                str(r["advance_date"]),
                r["employee_code"] or "",
                r["full_name"],
                float(r["amount"]),
                status_map.get(r["status"], r["status"]),
                r["reason"] or "",
                r["notes"] or ""
            ])

        auto_width(ws)
        filename = f"advances_{date_from or 'all'}_{date_to or 'all'}.xlsx"
        return create_excel_response(wb, filename)

# ── DASHBOARD ──
@app.get("/api/admin/dashboard",tags=["Admin"])
async def dashboard(u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        today=date.today(); tid=u["tenant_id"]
        total=await c.fetchval("SELECT COUNT(*) FROM users WHERE tenant_id=$1 AND is_active=TRUE AND role='employee'",tid)
        present=await c.fetchval("SELECT COUNT(DISTINCT user_id) FROM attendance WHERE tenant_id=$1 AND attendance_date=$2",tid,today)
        late=await c.fetchval("SELECT COUNT(DISTINCT user_id) FROM attendance WHERE tenant_id=$1 AND attendance_date=$2 AND status='late'",tid,today)
        sessions=await c.fetchval("SELECT COUNT(*) FROM attendance WHERE tenant_id=$1 AND attendance_date=$2",tid,today)
        records=await c.fetch(
            """SELECT u.full_name,u.employee_code,d.name as dept_name,
                      a.check_in_time,a.check_out_time,a.status,a.late_minutes,a.work_hours,a.session_number
               FROM attendance a JOIN users u ON a.user_id=u.id LEFT JOIN departments d ON u.department_id=d.id
               WHERE a.tenant_id=$1 AND a.attendance_date=$2 ORDER BY a.check_in_time DESC LIMIT 20""",tid,today)
        return {
            "stats":{"total_employees":total,"present_today":present,"absent_today":max(0,total-present),
                     "late_today":late,"total_sessions":sessions,
                     "attendance_rate":round(present/total*100 if total else 0,1)},
            "recent_records":[{"full_name":r["full_name"],"employee_code":r["employee_code"],
                "dept_name":r["dept_name"],"session_number":r["session_number"],
                "check_in_time":r["check_in_time"].isoformat() if r["check_in_time"] else None,
                "check_out_time":r["check_out_time"].isoformat() if r["check_out_time"] else None,
                "status":r["status"],"late_minutes":r["late_minutes"],
                "work_hours":float(r["work_hours"]) if r["work_hours"] else None} for r in records]}

@app.get("/api/admin/employees",tags=["Admin"])
async def employees_today(u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        emps=await c.fetch(
            "SELECT u.id,u.full_name,u.email,u.username,u.department_id,u.job_title,u.employee_code,u.role,u.hourly_rate,u.base_salary,u.salary_type,d.name as dept_name FROM users u LEFT JOIN departments d ON u.department_id=d.id WHERE u.tenant_id=$1 AND u.is_active=TRUE ORDER BY u.full_name",
            u["tenant_id"])
        result=[]
        for emp in emps:
            sessions=await c.fetch(
                "SELECT check_in_time,check_out_time,status,work_hours,session_number FROM attendance WHERE user_id=$1 AND attendance_date=$2 ORDER BY session_number",
                emp["id"],date.today())
            total_h=sum(float(r["work_hours"] or 0) for r in sessions)
            result.append({"id":s(emp["id"]),"full_name":emp["full_name"],"email":emp.get("email",""),"username":emp["username"],
                "dept_name":emp["dept_name"],"job_title":emp["job_title"],
                "employee_code":emp["employee_code"],"role":emp["role"],
                "hourly_rate":float(emp["hourly_rate"] or 0),"base_salary":float(emp["base_salary"] or 0),
                "salary_type":emp["salary_type"],
                "today":{"sessions":len(sessions),"total_hours":round(total_h,2),
                    "has_open":any(not r["check_out_time"] for r in sessions if r["check_in_time"]),
                    "last_status":sessions[-1]["status"] if sessions else None,
                    "first_checkin":sessions[0]["check_in_time"].isoformat() if sessions and sessions[0]["check_in_time"] else None}})
        return result

# ── REPORTS ──
@app.get("/api/reports/attendance",tags=["Reports"])
async def att_report(date_from:date=Query(...),date_to:date=Query(...),user_id:Optional[str]=None,u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        q="SELECT a.*,u.full_name,u.employee_code,u.salary_type,u.base_salary,d.name as dept_name FROM attendance a JOIN users u ON a.user_id=u.id LEFT JOIN departments d ON u.department_id=d.id WHERE a.tenant_id=$1 AND a.attendance_date BETWEEN $2 AND $3"
        params=[u["tenant_id"],date_from,date_to]
        if user_id: params.append(user_id); q+=f" AND a.user_id=${len(params)}"
        q+=" ORDER BY a.attendance_date DESC,u.full_name,a.session_number"
        rows=await c.fetch(q,*params)
        return [{"id":s(r["id"]),"date":str(r["attendance_date"]),"full_name":r["full_name"],
            "employee_code":r["employee_code"],"dept_name":r["dept_name"],
            "session_number":r["session_number"],
            "check_in":r["check_in_time"].isoformat() if r["check_in_time"] else None,
            "check_out":r["check_out_time"].isoformat() if r["check_out_time"] else None,
            "work_hours":float(r["work_hours"] or 0),"status":r["status"],
            "late_minutes":r["late_minutes"],"is_remote":r["is_remote"],
            "salary_type":r["salary_type"],"base_salary":float(r["base_salary"] or 0),
            "is_manual":r["is_manual"]} for r in rows]

# ── Export Attendance Report to Excel ──
@app.get("/api/reports/attendance/export",tags=["Reports"])
async def export_att_report(
    date_from:date=Query(...),
    date_to:date=Query(...),
    user_id:Optional[str]=None,
    u=Depends(admin_only)
):
    """تصدير تقرير الحضور إلى Excel"""
    async with db_pool.acquire() as c:
        q="SELECT a.*,u.full_name,u.employee_code,u.salary_type,u.base_salary,d.name as dept_name FROM attendance a JOIN users u ON a.user_id=u.id LEFT JOIN departments d ON u.department_id=d.id WHERE a.tenant_id=$1 AND a.attendance_date BETWEEN $2 AND $3"
        params=[u["tenant_id"],date_from,date_to]
        if user_id: params.append(user_id); q+=f" AND a.user_id=${len(params)}"
        q+=" ORDER BY a.attendance_date DESC,u.full_name,a.session_number"
        rows=await c.fetch(q,*params)

        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير الحضور"
        ws.sheet_view.rightToLeft = True

        headers = ["التاريخ", "الموظف", "الكود", "القسم", "جلسة", "الدخول", "الخروج", "الساعات", "الحالة", "تأخير", "نوع"]
        ws.append(headers)
        style_header(ws, 1, len(headers))

        status_map = {'present':'في الوقت','late':'متأخر','absent':'غائب','half_day':'نصف يوم','flexible':'مرن'}
        for r in rows:
            ws.append([
                str(r["attendance_date"]),
                r["full_name"],
                r["employee_code"] or "",
                r["dept_name"] or "",
                r["session_number"],
                r["check_in_time"].isoformat() if r["check_in_time"] else "",
                r["check_out_time"].isoformat() if r["check_out_time"] else "",
                float(r["work_hours"] or 0),
                status_map.get(r["status"], r["status"]),
                r["late_minutes"] or 0,
                "يدوي" if r["is_manual"] else "نظام"
            ])

        auto_width(ws)
        filename = f"attendance_report_{date_from}_{date_to}.xlsx"
        return create_excel_response(wb, filename)

@app.get("/api/reports/summary",tags=["Reports"])
async def summary_report(date_from:date=Query(...),date_to:date=Query(...),u=Depends(admin_only)):
    """ملخص الرواتب — يشمل المكافآت والخصومات من التسويات اليومية + الإضافية"""
    async with db_pool.acquire() as c:
        emps=await c.fetch(
            "SELECT u.id,u.full_name,u.employee_code,u.hourly_rate,u.base_salary,u.salary_type,d.name as dept_name FROM users u LEFT JOIN departments d ON u.department_id=d.id WHERE u.tenant_id=$1 AND u.is_active=TRUE AND u.role='employee' ORDER BY u.full_name",
            u["tenant_id"])
        result=[]
        for emp in emps:
            att=await c.fetch("SELECT attendance_date,status,late_minutes,work_hours FROM attendance WHERE user_id=$1 AND attendance_date BETWEEN $2 AND $3",emp["id"],date_from,date_to)

            adj_rows = await c.fetch(
                "SELECT adjustment_date,bonus,deduction,reason FROM daily_adjustments WHERE user_id=$1 AND adjustment_date BETWEEN $2 AND $3 ORDER BY adjustment_date",
                emp["id"], date_from, date_to)

            total_bonus = sum(float(r["bonus"] or 0) for r in adj_rows)
            total_ded = sum(float(r["deduction"] or 0) for r in adj_rows)

            adv=await c.fetchval("SELECT COALESCE(SUM(amount),0) FROM advances WHERE user_id=$1 AND advance_date BETWEEN $2 AND $3 AND status='approved'",emp["id"],date_from,date_to)
            days=len(set(str(r["attendance_date"]) for r in att))
            total_h=sum(float(r["work_hours"] or 0) for r in att)
            late_d=sum(1 for r in att if r["status"]=="late")

            if emp["salary_type"] == "fixed":
                daily_rate = float(emp["base_salary"] or 0) / 30
                gross = round(daily_rate * days, 2)
                rate_val = float(emp["base_salary"] or 0)
            else:
                gross = round(total_h * float(emp["hourly_rate"] or 0), 2)
                rate_val = float(emp["hourly_rate"] or 0)

            adv_t=float(adv or 0)
            net=round(gross+total_bonus-total_ded-adv_t,2)
            result.append({
                "id":s(emp["id"]),"full_name":emp["full_name"],"employee_code":emp["employee_code"],
                "dept_name":emp["dept_name"],"salary_type":emp["salary_type"],
                "hourly_rate":rate_val,"base_salary":float(emp["base_salary"] or 0),
                "work_days":days,"total_hours":round(total_h,2),"late_days":late_d,
                "gross_salary":gross,
                "total_bonus":total_bonus,
                "total_deduction":total_ded,
                "bonus":total_bonus,
                "deduction":total_ded,
                "advances":adv_t,"net_salary":net,
                "adjustments_detail":[{
                    "date":str(r["adjustment_date"]),
                    "bonus":float(r["bonus"] or 0),
                    "deduction":float(r["deduction"] or 0),
                    "reason":r["reason"]
                } for r in adj_rows]
            })
        return result

# ── Export Salary Summary to Excel ──
@app.get("/api/reports/summary/export",tags=["Reports"])
async def export_summary_report(
    date_from:date=Query(...),
    date_to:date=Query(...),
    u=Depends(admin_only)
):
    """تصدير ملخص الرواتب إلى Excel"""
    async with db_pool.acquire() as c:
        emps=await c.fetch(
            "SELECT u.id,u.full_name,u.employee_code,u.hourly_rate,u.base_salary,u.salary_type,d.name as dept_name FROM users u LEFT JOIN departments d ON u.department_id=d.id WHERE u.tenant_id=$1 AND u.is_active=TRUE AND u.role='employee' ORDER BY u.full_name",
            u["tenant_id"])

        wb = Workbook()
        ws = wb.active
        ws.title = "ملخص الرواتب"
        ws.sheet_view.rightToLeft = True

        headers = ["الموظف", "الكود", "القسم", "نوع الراتب", "الأساسي", "أيام", "ساعات", 
                   "إجمالي", "مكافآت التسويات", "خصومات التسويات", "مكافآت إضافية", "خصومات إضافية", "سلف", "الصافي"]
        ws.append(headers)
        style_header(ws, 1, len(headers))

        total_gross = total_bonus = total_ded = total_extra_bonus = total_extra_ded = total_adv = total_net = 0

        for emp in emps:
            att=await c.fetch("SELECT attendance_date,status,late_minutes,work_hours FROM attendance WHERE user_id=$1 AND attendance_date BETWEEN $2 AND $3",emp["id"],date_from,date_to)
            adj_rows = await c.fetch(
                "SELECT bonus,deduction FROM daily_adjustments WHERE user_id=$1 AND adjustment_date BETWEEN $2 AND $3",
                emp["id"], date_from, date_to)

            total_adj_bonus = sum(float(r["bonus"] or 0) for r in adj_rows)
            total_adj_ded = sum(float(r["deduction"] or 0) for r in adj_rows)

            adv=await c.fetchval("SELECT COALESCE(SUM(amount),0) FROM advances WHERE user_id=$1 AND advance_date BETWEEN $2 AND $3 AND status='approved'",emp["id"],date_from,date_to)
            days=len(set(str(r["attendance_date"]) for r in att))
            total_h=sum(float(r["work_hours"] or 0) for r in att)

            if emp["salary_type"] == "fixed":
                daily_rate = float(emp["base_salary"] or 0) / 30
                gross = round(daily_rate * days, 2)
            else:
                gross = round(total_h * float(emp["hourly_rate"] or 0), 2)

            adv_t=float(adv or 0)
            net=round(gross+total_adj_bonus-total_adj_ded-adv_t,2)

            total_gross += gross
            total_bonus += total_adj_bonus
            total_ded += total_adj_ded
            total_adv += adv_t
            total_net += net

            ws.append([
                emp["full_name"],
                emp["employee_code"] or "",
                emp["dept_name"] or "",
                "ثابت" if emp["salary_type"]=="fixed" else "بالساعة",
                float(emp["base_salary"] or 0),
                days,
                round(total_h, 2),
                gross,
                total_adj_bonus,
                total_adj_ded,
                0,  # extra_bonus placeholder
                0,  # extra_deduction placeholder
                adv_t,
                net
            ])

        # Add totals row
        ws.append(["", "", "", "", "", "", "الإجمالي", total_gross, total_bonus, total_ded, 0, 0, total_adv, total_net])
        last_row = ws.max_row
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=last_row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")

        auto_width(ws)
        filename = f"salary_summary_{date_from}_{date_to}.xlsx"
        return create_excel_response(wb, filename)

@app.post("/api/salary/calculate",tags=["Salary"])
async def calc_salary(r: SalaryCalcReq,u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        emp=await c.fetchrow("SELECT u.*,d.name as dept_name FROM users u LEFT JOIN departments d ON u.department_id=d.id WHERE u.id=$1 AND u.tenant_id=$2",r.user_id,u["tenant_id"])
        if not emp: raise HTTPException(404,"الموظف غير موجود")
        att=await c.fetch("SELECT attendance_date,session_number,check_in_time,check_out_time,status,late_minutes,work_hours FROM attendance WHERE user_id=$1 AND attendance_date BETWEEN $2 AND $3 ORDER BY attendance_date,session_number",r.user_id,r.date_from,r.date_to)
        adj=await c.fetchrow("SELECT COALESCE(SUM(bonus),0) as tb,COALESCE(SUM(deduction),0) as td FROM daily_adjustments WHERE user_id=$1 AND adjustment_date BETWEEN $2 AND $3",r.user_id,r.date_from,r.date_to)
        adv_rows=await c.fetch("SELECT * FROM advances WHERE user_id=$1 AND advance_date BETWEEN $2 AND $3 AND status='approved' ORDER BY advance_date",r.user_id,r.date_from,r.date_to)
        days=len(set(str(a["attendance_date"]) for a in att))
        total_h=sum(float(a["work_hours"] or 0) for a in att)

        if emp["salary_type"] == "fixed":
            daily_rate = float(emp["base_salary"] or 0) / 30
            gross = round(daily_rate * days, 2)
            rate_display = float(emp["base_salary"] or 0)
        else:
            rate = float(emp["hourly_rate"] or 0)
            gross = round(total_h * rate, 2)
            rate_display = rate

        adj_bonus=float(adj["tb"] or 0)
        adj_ded=float(adj["td"] or 0)
        total_bonus = adj_bonus + r.extra_bonus
        total_ded = adj_ded + r.extra_deduction
        adv_total=sum(float(a["amount"]) for a in adv_rows)
        net=round(gross+total_bonus-total_ded-adv_total,2)
        return {
            "employee":{"id":s(emp["id"]),"full_name":emp["full_name"],"employee_code":emp["employee_code"],"dept_name":emp["dept_name"],"hourly_rate":rate_display,"base_salary":float(emp["base_salary"] or 0),"salary_type":emp["salary_type"]},
            "period":{"from":str(r.date_from),"to":str(r.date_to)},
            "summary":{"work_days":days,"total_sessions":len(att),"total_hours":round(total_h,2),"hourly_rate":rate_display,"base_salary":float(emp["base_salary"] or 0),"salary_type":emp["salary_type"],"gross_salary":gross,
                        "adj_bonus":adj_bonus,"adj_deduction":adj_ded,
                        "extra_bonus":r.extra_bonus,"extra_deduction":r.extra_deduction,
                        "bonus":total_bonus,"deduction":total_ded,
                        "advances":adv_total,"net_salary":net},
            "daily_records":[{"date":str(a["attendance_date"]),"session":a["session_number"],"check_in":a["check_in_time"].isoformat() if a["check_in_time"] else None,"check_out":a["check_out_time"].isoformat() if a["check_out_time"] else None,"hours":float(a["work_hours"] or 0),"status":a["status"],"late":a["late_minutes"]} for a in att],
            "advances":[{"date":str(a["advance_date"]),"amount":float(a["amount"]),"reason":a["reason"]} for a in adv_rows]}

@app.post("/api/salary/save",tags=["Salary"])
async def save_salary(r: SalaryCalcReq,u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        emp=await c.fetchrow("SELECT hourly_rate,base_salary,salary_type FROM users WHERE id=$1 AND tenant_id=$2",r.user_id,u["tenant_id"])
        if not emp: raise HTTPException(404,"الموظف غير موجود")
        att=await c.fetch("SELECT work_hours,attendance_date FROM attendance WHERE user_id=$1 AND attendance_date BETWEEN $2 AND $3",r.user_id,r.date_from,r.date_to)
        adj=await c.fetchrow("SELECT COALESCE(SUM(bonus),0) as tb,COALESCE(SUM(deduction),0) as td FROM daily_adjustments WHERE user_id=$1 AND adjustment_date BETWEEN $2 AND $3",r.user_id,r.date_from,r.date_to)
        adv=await c.fetchval("SELECT COALESCE(SUM(amount),0) FROM advances WHERE user_id=$1 AND advance_date BETWEEN $2 AND $3 AND status='approved'",r.user_id,r.date_from,r.date_to)
        days=len(set(str(a["attendance_date"]) for a in att))
        total_h=sum(float(a["work_hours"] or 0) for a in att)

        if emp["salary_type"] == "fixed":
            daily_rate = float(emp["base_salary"] or 0) / 30
            gross = round(daily_rate * days, 2)
            rate_display = float(emp["base_salary"] or 0)
        else:
            rate = float(emp["hourly_rate"] or 0)
            gross = round(total_h * rate, 2)
            rate_display = rate

        bonus=float(adj["tb"] or 0)+r.extra_bonus
        ded=float(adj["td"] or 0)+r.extra_deduction
        adv_t=float(adv or 0)
        net=round(gross+bonus-ded-adv_t,2)
        row=await c.fetchrow(
            "INSERT INTO salary_reports (tenant_id,user_id,report_name,report_from,report_to,total_days,total_sessions,total_hours,hourly_rate,base_salary,salary_type,gross_salary,total_bonus,total_deduction,total_advances,net_salary,notes,created_by) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17,$18) RETURNING id",
            u["tenant_id"],r.user_id,r.report_name,r.date_from,r.date_to,days,len(att),total_h,rate_display,float(emp["base_salary"] or 0),emp["salary_type"],gross,bonus,ded,adv_t,net,r.notes,u["user_id"])
        return {"success":True,"message":"تم حفظ التقرير","net_salary":net,"id":s(row["id"])}

@app.get("/api/salary/reports",tags=["Salary"])
async def saved_reports(u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        rows=await c.fetch(
            "SELECT sr.*,u.full_name,u.employee_code FROM salary_reports sr JOIN users u ON sr.user_id=u.id WHERE sr.tenant_id=$1 ORDER BY sr.created_at DESC LIMIT 100",
            u["tenant_id"])
        return [dict(r)|{"id":s(r["id"]),"user_id":s(r["user_id"]),
            "total_hours":float(r["total_hours"]),"gross_salary":float(r["gross_salary"]),
            "net_salary":float(r["net_salary"]),"total_bonus":float(r["total_bonus"]),
            "total_deduction":float(r["total_deduction"]),"total_advances":float(r["total_advances"]),
            "base_salary":float(r["base_salary"] or 0),"salary_type":r["salary_type"],
            "report_from":str(r["report_from"]),"report_to":str(r["report_to"]),
            "created_at":r["created_at"].isoformat()} for r in rows]

# ── Export Saved Reports to Excel ──
@app.get("/api/salary/reports/export",tags=["Salary"])
async def export_saved_reports(u=Depends(admin_only)):
    """تصدير التقارير المحفوظة إلى Excel"""
    async with db_pool.acquire() as c:
        rows=await c.fetch(
            "SELECT sr.*,u.full_name,u.employee_code FROM salary_reports sr JOIN users u ON sr.user_id=u.id WHERE sr.tenant_id=$1 ORDER BY sr.created_at DESC",
            u["tenant_id"])

        wb = Workbook()
        ws = wb.active
        ws.title = "التقارير المحفوظة"
        ws.sheet_view.rightToLeft = True

        headers = ["الاسم", "الموظف", "الكود", "الفترة", "نوع الراتب", "الأساسي", "أيام", "ساعات", 
                   "إجمالي", "مكافآت", "خصومات", "سلف", "الصافي", "التاريخ"]
        ws.append(headers)
        style_header(ws, 1, len(headers))

        for r in rows:
            ws.append([
                r["report_name"] or "",
                r["full_name"],
                r["employee_code"] or "",
                f"{r['report_from']} → {r['report_to']}",
                "ثابت" if r["salary_type"]=="fixed" else "بالساعة",
                float(r["base_salary"] or 0),
                r["total_days"],
                float(r["total_hours"]),
                float(r["gross_salary"]),
                float(r["total_bonus"]),
                float(r["total_deduction"]),
                float(r["total_advances"]),
                float(r["net_salary"]),
                r["created_at"].strftime("%Y-%m-%d")
            ])

        auto_width(ws)
        filename = f"saved_reports_{date.today()}.xlsx"
        return create_excel_response(wb, filename)

# ── SALARY REPORTS CRUD ──
@app.put("/api/salary/reports/{rpt_id}",tags=["Salary"])
async def update_salary_report(rpt_id: str, r: UpdateSalaryReportReq, u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        rpt = await c.fetchrow(
            "SELECT id FROM salary_reports WHERE id=$1 AND tenant_id=$2",
            rpt_id, u["tenant_id"])
        if not rpt: raise HTTPException(404, "التقرير غير موجود")
        data = r.model_dump(exclude_none=True)
        if not data: raise HTTPException(400, "لا توجد بيانات")
        fields = [f"{k}=${i+1}" for i, k in enumerate(data.keys())]
        vals = list(data.values()) + [rpt_id]
        await c.execute(f"UPDATE salary_reports SET {','.join(fields)} WHERE id=${len(vals)}", *vals)
        return {"success": True, "message": "تم تحديث التقرير"}

@app.delete("/api/salary/reports/{rpt_id}",tags=["Salary"])
async def delete_salary_report(rpt_id: str, u=Depends(admin_only)):
    async with db_pool.acquire() as c:
        result = await c.execute(
            "DELETE FROM salary_reports WHERE id=$1 AND tenant_id=$2",
            rpt_id, u["tenant_id"])
        if result == "DELETE 0": raise HTTPException(404, "التقرير غير موجود")
        return {"success": True, "message": "تم حذف التقرير"}


# ── Run for Render / Local ──
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=PORT)
